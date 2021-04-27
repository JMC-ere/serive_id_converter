from openpyxl import load_workbook


def read_xl(logger):

    try:
        service_ids_list = []

        # 엑셀 파일 읽어오기
        service_id_xl = load_workbook(filename="C:/Users/정민채/OneDrive/바탕 화면/Files/test.xlsx", data_only=True)
        # 시트 이름 구하기
        sheet_name = service_id_xl.sheetnames[0]

        service_sheet = service_id_xl[sheet_name]

        for service_id in service_sheet['A']:
            service_ids_list.append(str(service_id.value))

        logger.info('TOTAL SERVICE ID COUNT :' + str(len(service_ids_list)))

        return service_ids_list

    except Exception as get_err:

        logger.error('GET SERVICE ID ERROR :', get_err)
        exit()
